
import numpy as np
import socket
import random
import time
import threading
import argparse
import sqlite3
from datetime import datetime
from datetime import date,timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
import array as arr


# List of holiday is as per year 2021:
list_of_holidays = ["JAN 14","JAN 26","APR 02","APR 13","APR 25","MAY 14","MAY 26","JUL 21","AUG 15","AUG 19","SEP 10","OCT 02","OCT 15","OCT 19","NOV 01","NOV 04","NOV 19","DEC 25"]     

dict_month_index={"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}

dict_month_selector={1:"JAN", 2:"FEB",3:"MAR",4:"APR",5:"MAY",6:"JUN",7:"JUL",8:"AUG",9:"SEP",10:"OCT",11:"NOV",12:"DEC"}

dict_employee_twinklet_id={"Devamma":7, "Lakshmamma":6, "Manjula N":3, "Venkat Ramanamma":4}

def monthly_attendance_excel_headers(ws,entered_year,entered_month):
    ws['N1'] = "Form XVI"
    n1 = ws['N1']
    n1.font = Font(size=17, bold=True)
    ws.row_dimensions[1].height = 17

    ws['C2'] = "Register of Attendance for the month of "+entered_month+" "+entered_year
    c2 = ws['C2']
    c2.font = Font(size=22, bold=True)
    ws.row_dimensions[2].height = 22
    
    ws['A4'] = "Name of the firm: IISc, ECE"
    a4 = ws['A4']
    a4.font = Font(size=14, bold=True)

    ws['J4'] = "Nature of Work: Un-skilled"
    j4 = ws['J4']
    j4.font = Font(size=14, bold=True)

    ws['X4'] = "Place: IISc, Bengaluru"
    x4 = ws['X4']
    x4.font = Font(size=14, bold=True)

    ws['A5'] = "Names"
    a5 = ws['A5']
    a5.font = Font(size=12, bold=True)
    a5.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"),
                       bottom=Side(border_style="thin"))
    a5.alignment = Alignment(horizontal="center")

    ws['A6'] = "Devamma"            #employee1
    ws['A7'] = "Lakshmamma"         #employee2
    ws['A8'] = "Manjula N"          #employee3
    ws['A9'] = "Venkat Ramanamma"   #employee4

    ws.column_dimensions['A'].width = 19

    i = 6
    while i <= 9:
        names = ws.cell(row=i, column=1)
        names.alignment = Alignment(horizontal="general", wrap_text=False, shrink_to_fit=False, shrinkToFit=False)
        names.font = Font(condense=False, extend=True)
        i += 1

    i = 1
    while i <= 34:
        cell6 = ws.cell(row=6, column=i)
        cell6.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell6.alignment = Alignment(horizontal="center")

        cell7 = ws.cell(row=7, column=i)
        cell7.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell7.alignment = Alignment(horizontal="center")

        cell8 = ws.cell(row=8, column=i)
        cell8.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell8.alignment = Alignment(horizontal="center")

        cell9 = ws.cell(row=9, column=i)
        cell9.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        cell9.alignment = Alignment(horizontal="center")
        i += 1

    i = 2
    while i <= 34:
        ws.cell(row=5, column=i).value = i-1
        anum = ws.cell(row=5, column=i)
        anum.font = Font(size=12, bold=True)
        anum.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        anum.alignment = Alignment(horizontal="center")
        i += 1

    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 3
    ws.column_dimensions['I'].width = 3
    ws.column_dimensions['J'].width = 3
    ws.column_dimensions['K'].width = 3
    ws.column_dimensions['L'].width = 3
    ws.column_dimensions['M'].width = 3
    ws.column_dimensions['N'].width = 3
    ws.column_dimensions['O'].width = 3
    ws.column_dimensions['P'].width = 3
    ws.column_dimensions['Q'].width = 3
    ws.column_dimensions['R'].width = 3
    ws.column_dimensions['S'].width = 3
    ws.column_dimensions['T'].width = 3
    ws.column_dimensions['U'].width = 3
    ws.column_dimensions['V'].width = 3
    ws.column_dimensions['W'].width = 3
    ws.column_dimensions['X'].width = 3
    ws.column_dimensions['Y'].width = 3
    ws.column_dimensions['Z'].width = 3
    ws.column_dimensions['AA'].width = 3
    ws.column_dimensions['AB'].width = 3
    ws.column_dimensions['AC'].width = 3
    ws.column_dimensions['AD'].width = 3
    ws.column_dimensions['AE'].width = 3
    ws.column_dimensions['AF'].width = 3

    i = 6
    j = 2
    while i <= 9:
        while j <= 34:
            ws.cell(row=i, column=j).value = " "
            anum = ws.cell(row=i, column=j)
            anum.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            j += 1
        i += 1

    ws.cell(row=5, column=33).value = "WO"
    ag5 = ws.cell(row=5, column=33)
    ag5.font = Font(size=12, bold=True)
    ag5.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ag5.alignment = Alignment(horizontal="center")
    ws.column_dimensions['AG'].width = 5

    ws.cell(row=5, column=34).value = "TD"
    ah5 = ws.cell(row=5, column=34)
    ah5.font = Font(size=12, bold=True)
    ah5.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ah5.alignment = Alignment(horizontal="center")
    ws.column_dimensions['AH'].width = 5

    ws['B12'] = "P: Present"
    b12 = ws['B12']
    b12.font = Font(size=11)
    b12.border = Border(left=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['B13'] = "A: Absent"
    b13 = ws['B13']
    b13.font = Font(size=11)
    b13.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['H12'] = "WO: Work off"
    h12 = ws['H12']
    h12.font = Font(size=11)
    h12.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L12'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['H13'] = "TD: Total days"
    h13 = ws['H13']
    h13.font = Font(size=11)
    h13.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L13'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['B15'].fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
    ws['C15'] = ": Sunday"
    c15 = ws['C15']
    c15.font = Font(size=11)
    ws['B15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['H15'].fill = PatternFill(fill_type='solid', start_color='000000FF', end_color='000000FF')
    ws['I15'] = ": Institute Holiday"
    i15 = ws['I15']
    i15.font = Font(size=11)
    ws['H15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['M15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['N15'].border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['B17'] = "DEBIT HEAD: Part(1A) WE.01-0402-0000-44-433"
    b17 = ws['B17']
    b17.font = Font(size=11, bold=True)

    ws['B26'] = "Supervisor's Signature:"
    b26 = ws['B26']
    b26.font = Font(size=11, bold=True)

    ws['T26'] = "Chairman's Signature:"
    t26 = ws['T26']
    t26.font = Font(size=11, bold=True)
    
    
def monthly_attendance_find_sundays(year, month):
    year = int(year)
    month = dict_month_index[month]
    d = date(year,month,1)
    d += timedelta(days=6-d.weekday())      #First Sunday
    while d.year == year:
        if d.month == month:
            yield d
            d += timedelta(days = 7)
            #print("d:",d)

        if d.month != month:
            break


def monthly_attendance_sundays(wb,ws,year,month):
    for sundays in monthly_attendance_find_sundays(year, month):
        #print(sundays)
        sundays = str(sundays)

        data_split = sundays.split("-")
        data_split_date = int(data_split[2])
            
        i = 6
        while i <= 9:
            ws.cell(row=i, column=data_split_date + 1).fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
            ws.cell(row=i, column=data_split_date + 1).value = " "
            i += 1
    wb.save(month+"-"+year+".xlsx")


def monthly_attendance_institute_holidays(wb,ws,year,month):
    for items in list_of_holidays:
        mon_date = items
        mon_date = mon_date.split()
        list_of_hol_mon = mon_date[0]
        list_of_hol_date = int(mon_date[1])
        
        if list_of_hol_mon == month:
            #print("list_of_hol_date:",list_of_hol_date)
            if list_of_hol_date:
                i = 6
                while i <= 9:
                    ws.cell(row=i, column=list_of_hol_date+1).fill = PatternFill(fill_type='solid', start_color='000000FF', end_color='000000FF')
                    ws.cell(row=i, column=list_of_hol_date+1).value = " "
                        # ws.cell(row=6, column=i).value = "IH"
                    i += 1

    wb.save(month+"-"+year+".xlsx")



total_number_of_months = 0

def query_database_excel_headers(ws, employee_name, start_year, start_month, end_year, end_month):
    global total_number_of_months
    entered_year = start_year
    entered_month = start_month

    ws['B2'] = "Register of Attendance for the month of "+ start_month+ " "+ str(start_year)+ "-"+ end_month+ " "+ str(end_year)
    b2 = ws['B2']
    b2.font = Font(size=22,bold=True)
    ws.row_dimensions[2].height = 22
    
    ws['A4'] = "Name of the firm: IISc, ECE"
    a4 = ws['A4']
    a4.font = Font(size=14, bold=True)

    ws['J4'] = "Nature of Work: Un-skilled"
    j4 = ws['J4']
    j4.font = Font(size=14, bold=True)

    ws['X4'] = "Place: IISc, Bengaluru"
    x4 = ws['X4']
    x4.font = Font(size=14, bold=True)

    ws['A5'] = "Name of employee:"
    a5 = ws['A5']
    a5.font = Font(size=14, bold=True)

    ws['D5'] = employee_name
    d5 = ws['D5']
    d5.font = Font(size=14)

    ws['A6'] = "Month"
    a6 = ws['A6']
    a6.font = Font(size=12, bold=True)
    a6.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"),
                       bottom=Side(border_style="thin"))
    a6.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 17

    i = 2
    while i <= 32:
        ws.cell(row=6, column=i).value = i-1
        anum = ws.cell(row=6, column=i)
        anum.font = Font(size=12, bold=True)
        anum.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        anum.alignment = Alignment(horizontal="center")
        i += 1

    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 3
    ws.column_dimensions['I'].width = 3
    ws.column_dimensions['J'].width = 3
    ws.column_dimensions['K'].width = 3
    ws.column_dimensions['L'].width = 3
    ws.column_dimensions['M'].width = 3
    ws.column_dimensions['N'].width = 3
    ws.column_dimensions['O'].width = 3
    ws.column_dimensions['P'].width = 3
    ws.column_dimensions['Q'].width = 3
    ws.column_dimensions['R'].width = 3
    ws.column_dimensions['S'].width = 3
    ws.column_dimensions['T'].width = 3
    ws.column_dimensions['U'].width = 3
    ws.column_dimensions['V'].width = 3
    ws.column_dimensions['W'].width = 3
    ws.column_dimensions['X'].width = 3
    ws.column_dimensions['Y'].width = 3
    ws.column_dimensions['Z'].width = 3
    ws.column_dimensions['AA'].width = 3
    ws.column_dimensions['AB'].width = 3
    ws.column_dimensions['AC'].width = 3
    ws.column_dimensions['AD'].width = 3
    ws.column_dimensions['AE'].width = 3
    ws.column_dimensions['AF'].width = 3

    ws.cell(row=6, column=33).value = "WO"
    ag6 = ws.cell(row=6, column=33)
    ag6.font = Font(size=12, bold=True)
    ag6.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ag6.alignment = Alignment(horizontal="center")
    ws.column_dimensions['AG'].width = 5

    ws.cell(row=6, column=34).value = "TD"
    ah6 = ws.cell(row=6, column=34)
    ah6.font = Font(size=12, bold=True)
    ah6.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ah6.alignment = Alignment(horizontal="center")
    ws.column_dimensions['AH'].width = 5

    start_month_index = dict_month_index[start_month]
    end_month_index = dict_month_index[end_month]
    entered_month_index = start_month_index
    i = 7
    while end_year >= entered_year:
        month = dict_month_selector[entered_month_index]
        month_entry = month+ " "+str(entered_year)
        ws.cell(row=i, column=1).value=month_entry
        month_entry_i = ws.cell(row=i, column=1)
        month_entry_i.alignment = Alignment(horizontal="center")
        month_entry_i.border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
       
        if end_year == entered_year and end_month_index == entered_month_index:
            break
        
        i+=1
        entered_month_index +=1

        if entered_month_index == 13:
            entered_month_index =1
            entered_year += 1

    total_number_of_months = i

    #Borders
    j = 7
    k = 2
    while j <= i:
        if k <= 35:
            ws.cell(row=j, column=k).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
            ws.cell(row=j, column=k).value = " "
            k += 1
        if k == 35:
            k = 2
            j += 1

    ws.cell(row=i+3, column=2).value = "P: Present"
    ws.cell(row=i+3, column=2).font = Font(size=11)
    ws.cell(row=i+3, column=2).border = Border(left=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=3).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=4).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=5).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws.cell(row=i+4, column=2).value = "A: Absent"
    ws.cell(row=i+4, column=2).font = Font(size=11)
    ws.cell(row=i+4, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=3).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=4).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=5).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws.cell(row=i+3, column=8).value = "WO: Work off"
    ws.cell(row=i+3, column=8).font = Font(size=11)
    ws.cell(row=i+3, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=9).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=10).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=11).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+3, column=12).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws.cell(row=i+4, column=8).value = "TD: Total days"
    ws.cell(row=i+4, column=8).font = Font(size=11)
    ws.cell(row=i+4, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=9).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=10).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=11).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+4, column=12).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws.cell(row=i+6, column=2).fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
    ws.cell(row=i+6, column=3).value = ": Sunday"
    ws.cell(row=i+6, column=3).font = Font(size=11)
    ws.cell(row=i+6, column=2).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=3).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=4).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=5).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))

    ws.cell(row=i+6, column=8).fill = PatternFill(fill_type='solid', start_color='000000FF', end_color='000000FF')
    ws.cell(row=i+6, column=9).value = ": Institute Holiday"
    ws.cell(row=i+6, column=9).font = Font(size=11)
    ws.cell(row=i+6, column=8).border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=9).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=10).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=11).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=12).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=13).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws.cell(row=i+6, column=14).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"))


def query_database_find_sundays(start_year, start_month, start_date, end_year, end_month, end_date):
    entered_year = int(start_year)
    entered_month = dict_month_index[start_month]
    entered_date = int(start_date)

#    end_year = int(end_year)
#    end_month = dict_month_index[end_month]
#    end_date = int(end_date)

    d = date(entered_year,entered_month,entered_date)
    d += timedelta(days=6-d.weekday())      #First Sunday
    while d.year <= end_year:
        yield d
        d += timedelta(days = 7)
#        print("d:",d)

        if d.year == end_year:
            if d.month == end_month:
                if d.day > end_date:
                    break


def query_database_sundays(wb, ws, employee_name, start_year, start_month, start_date, end_year, end_month, end_date):

    end_year = int(end_year)
    end_month = dict_month_index[end_month]
    end_date = int(end_date)
    prev_data_split_month = dict_month_index[start_month]
    i = 7
    
    for sundays in query_database_find_sundays(start_year, start_month, start_date, end_year, end_month, end_date):
        sundays = str(sundays)
#        print("sundays:",sundays)

        data_split = sundays.split("-")
        data_split_year = int(data_split[0])
        data_split_month = int(data_split[1])
        data_split_date = int(data_split[2])
            
        if prev_data_split_month != data_split_month:
            i +=1

        ws.cell(row=i, column=data_split_date + 1).fill = PatternFill(fill_type='solid', start_color='00FF0000', end_color='00FF0000')
        ws.cell(row=i, column=data_split_date + 1).value = " "
        
        prev_data_split_month = data_split_month

    wb.save(employee_name + ".xlsx")

            

def query_database_institute_holidays(wb, ws, employee_name, start_year, start_month, start_date, end_year, end_month, end_date):
    prev_data_split_month = start_month
    end_month_index = dict_month_index[end_month]
    end_date = int(end_date)
    i = 7
    for items in list_of_holidays:
        mon_date = items
        mon_date = mon_date.split()
        list_of_hol_mon = mon_date[0]
        hol_month_index = dict_month_index[list_of_hol_mon]
        list_of_hol_date = int(mon_date[1])

        
        if hol_month_index > end_month_index:
            break
        elif hol_month_index == end_month_index: 
            if list_of_hol_date > end_date:
                break
        
        if prev_data_split_month != list_of_hol_mon:
            i +=1

        ws.cell(row=i, column=list_of_hol_date+1).fill = PatternFill(fill_type='solid', start_color='000000FF', end_color='000000FF')
        ws.cell(row=i, column=list_of_hol_date+1).value = " "
        
        prev_data_split_month = list_of_hol_mon

    wb.save(employee_name+".xlsx")

