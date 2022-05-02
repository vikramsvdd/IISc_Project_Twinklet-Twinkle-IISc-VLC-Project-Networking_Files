from tkinter import *
import numpy as np
import sqlite3
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
import array as arr
from enum import Enum

xs, ys = 7, 10000;


def spreadsheet_display():
    wb = Workbook()
    sheet = wb.active
    
    count = np.load("checkpoint_delay.npy")
    local_count = int(np.copy(count))
    iter_count = int(local_count)
    print(iter_count)

    sheet["A1"] = "S.No"
    sheet["B1"] = "TwinkleT ID"
    sheet["C1"] = "TwinkleR ID"
    sheet["D1"] = "Date & time"
    sheet["E1"] = "Date Pack"
    sheet["F1"] = "Delay"
    sheet["G1"] = "Statistics"

    sheet.column_dimensions['A'].width=5
    sheet.column_dimensions['B'].width=10
    sheet.column_dimensions['C'].width=10
    sheet.column_dimensions['D'].width=20
    sheet.column_dimensions['E'].width=10
    sheet.column_dimensions['F'].width=10
    sheet.column_dimensions['G'].width=15

    i = 1
    while i <= 7:
        sheet.cell(row=1,column=i).border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"))
        i+=1

    if iter_count < ys:
        cursor = conn.execute("SELECT SRNO, TWINKLETID, TWINKLERID, DATETIME, DATA_PACK, DELAY, STATISTICS from TWINKLE_DELAY_STATS_DATA WHERE SRNO<:SRNO", {"SRNO": (iter_count + 1)})
    else:
        cursor = conn.execute("SELECT SRNO, TWINKLETID, TWINKLERID, DATETIME, DATA_PACK, DELAY, STATISTICS from TWINKLE_DELAY_STATS_DATA WHERE SRNO BETWEEN ? AND ?", ((iter_count - (ys - 2)), iter_count))



    rowx = 0
    for rowr in cursor:
        if rowr[1] == "--":
            sheet.row_dimensions[rowx+2].height=50
        for colx in range(xs):
            sheet.cell(row=rowx+2,column=colx+1).value=rowr[colx]
        rowx = rowx + 1
    count = local_count
    wb.save("TCP_Delay_Statistics.xls")
    print("File Generated!")



if __name__ == "__main__":
    conn = sqlite3.connect("Twinkle_Delay_Stats_Database.db")
    print("Opened Database successfully...")
    spreadsheet_display()
