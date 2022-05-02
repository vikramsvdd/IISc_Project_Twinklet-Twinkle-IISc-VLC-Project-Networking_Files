
import numpy as np
import socket
import random
import time
import threading
import argparse
import sqlite3
from datetime import datetime
from datetime import date,timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
import array as arr

from Spreadsheet_headers import *
import Spreadsheet_headers

xs, ys = 7, 10000;

def monthly_attendance_calculate_wo_td(ws):
    wo = 0
    td = 0

    entered_attendance = ""
    i = 6
    j = 2
    while i <= 9:
        while j <= 32:

            entered_attendance = ws.cell(row=i, column=j)
        
            if entered_attendance.value == 'P':
                td += 1
            elif entered_attendance.value == 'A':
                wo += 1

            j += 1

        ws.cell(row=i, column=33).value = wo
        ws.cell(row=i, column=34).value = td
        j = 2
        wo = 0
        td = 0
        i+=1
        

def monthly_attendance_sheet(conn, entered_year, entered_month):
    global count
    global data_base
    global local_count

    entered_month = entered_month.upper()
    count = np.load("checkpoint.npy")
    local_count = int(np.copy(count))
    iter_count = int(local_count)
    print(iter_count)

    wb = Workbook()
    ws = wb.active

    monthly_attendance_excel_headers(ws, entered_year, entered_month)

    ############# INSTITUTE HOLIDAYS ################
    monthly_attendance_institute_holidays(wb,ws,entered_year,entered_month)

    ############### Sundays ########################
    monthly_attendance_sundays(wb,ws,entered_year,entered_month)

    if iter_count < ys:
        cursor = conn.execute(
            "SELECT SRNO, TWINKLERID, FLASHPOINTID, FLASHPOS, DATETIME, DATA_PACK, REMARKS from TWINKLE_DATA WHERE SRNO<:SRNO",
            {"SRNO": (iter_count + 1)})
    else:
        cursor = conn.execute(
            "SELECT SRNO, TWINKLERID, FLASHPOINTID, FLASHPOS, DATETIME, DATA_PACK, REMARKS from TWINKLE DATA WHERE SRNO BETWEEN ? AND ?",
            ((iter_count - (yc - 2)), iter_count))

    rowx = 1
    attendance = 'A'
    employee1_attendance_flag = 0
    employee2_attendance_flag = 0
    employee3_attendance_flag = 0
    employee4_attendance_flag = 0
    prev_date = 32
    for row in cursor:
        detail_date = row[4]
        day = detail_date[0:3]
        month = detail_date[4:7].upper()
        date = detail_date[8:10]
        year = detail_date[20:24]

        date = int(date)
        twinklet_id = row[1]
        #print("date: ", date, "twinklet_id: ", twinklet_id)
        rowx += 1

        if year == entered_year:
            if month == entered_month:
                if date:
                    if twinklet_id == 7:
                        attendance = 'P'
                        employee1_attendance_flag = 1
                        ws.cell(row=6, column=date+1).value = attendance
                    if twinklet_id == 6:
                        attendance = 'P'
                        employee2_attendance_flag = 1
                        ws.cell(row=7, column=date+1).value = attendance
                    if twinklet_id == 3:
                        attendance = 'P'
                        employee3_attendance_flag = 1
                        ws.cell(row=8, column=date+1).value = attendance
                    if twinklet_id == 4:
                        attendance = 'P'
                        employee4_attendance_flag = 1
                        ws.cell(row=9, column=date+1).value = attendance
              
                if prev_date != date and employee1_attendance_flag == 0:
                    ws.cell(row=6, column=date+1).value = 'A'
                if prev_date != date and employee2_attendance_flag == 0:
                    ws.cell(row=7, column=date+1).value = 'A'
                if prev_date != date and employee3_attendance_flag == 0:
                    ws.cell(row=8, column=date+1).value = 'A'
                if prev_date != date and employee4_attendance_flag == 0:
                    ws.cell(row=9, column=date+1).value = 'A'

                prev_date = date

                employee1_attendance_flag = 0
                employee2_attendance_flag = 0
                employee3_attendance_flag = 0
                employee4_attendance_flag = 0

    monthly_attendance_calculate_wo_td(ws)

    wb.save(entered_month+"-"+entered_year+".xlsx")

    count = local_count
    print("Generated file...")

def query_database_calculate_wo_td(ws):
    total_number_of_months = Spreadsheet_headers.total_number_of_months
    wo = 0
    td = 0

    entered_attendance = ""
    i = 7
    j = 2
    while i <= total_number_of_months:
        while j <= 32:
            entered_attendance = ws.cell(row=i, column=j)
        
            if entered_attendance.value == 'P':
                td += 1
            elif entered_attendance.value == 'A':
                wo += 1
            j+=1
        ws.cell(row=i, column=33).value = wo
        ws.cell(row=i, column=34).value = td
        j = 2
        wo = 0
        td = 0
        i+=1


def query_database(conn, employee_name, start_year, start_month, start_date, end_year, end_month, end_date):
    global count
    global local_count
    global date_base

    count = np.load("checkpoint.npy")
    local_count = int(np.copy(count))
    iter_count = int(local_count)
    print(iter_count)

    start_year = int(start_year)
    start_month = start_month.upper()
    start_month_index = dict_month_index[start_month]
    start_date = int(start_date)


    end_year = int(end_year)
    end_month = end_month.upper()
    end_month_index = dict_month_index[end_month]
    end_date = int(end_date)

    entered_year = start_year
    entered_month_index = start_month_index
    entered_date = start_date

    start_month_flag = 0
    start_date_flag = 0
    prev_date = 0
    attendance = 'A'
    employee_attendance_flag = 0

    employee_twinklet_id = dict_employee_twinklet_id[employee_name]

    i = 7

    wb = Workbook()
    ws = wb.active

    query_database_excel_headers(ws, employee_name, start_year, start_month, end_year, end_month)

#    ############# INSTITUTE HOLIDAYS ################
    query_database_institute_holidays(wb, ws, employee_name, start_year, start_month, start_date, end_year, end_month, end_date)

#    ############### Sundays ########################
    query_database_sundays(wb, ws, employee_name, start_year, start_month, start_date, end_year, end_month, end_date)


    if iter_count < ys:
        cursor = conn.execute(
            "SELECT SRNO, TWINKLERID, FLASHPOINTID, FLASHPOS, DATETIME, DATA_PACK, REMARKS from TWINKLE_DATA WHERE SRNO<:SRNO",
            {"SRNO": (iter_count + 1)}) #What is this??
    else:
        cursor = conn.execute(
            "SELECT SRNO, TWINKLERID, FLASHPOINTID, FLASHPOS, DATETIME, DATA_PACK, REMARKS from TWINKLE DATA WHERE SRNO BETWEEN ? AND ?",
            ((iter_count - (yc - 2)), iter_count))

    rowx = 1
    for row in cursor:
        twinklet_id = row[1]
        detail_date = row[4]
        day = detail_date[0:3]
        month = detail_date[4:7].upper()
        month_index = dict_month_index[month]
        date = int(detail_date[8:10])
        year = int(detail_date[20:24])

        rowx += 1

        if entered_year >= start_year and entered_year<= end_year:
            if entered_month_index == start_month_index or start_month_flag == 1:
                start_month_flag = 1
                if entered_date == start_date or start_date_flag == 1:
                    start_date_flag = 1
                    
                    if prev_date != date:
                        entered_date = prev_date
                    if month_index != entered_month_index:
                        entered_month_index += 1
                        i += 1
                        if entered_month_index >= 12:
                            entered_month_index = 1
                    if entered_month_index >= 12:
                        entered_year += 1

                    if twinklet_id == employee_twinklet_id:
#                        print(detail_date)
                        attendance = 'P'
                        employee_attendance_flag = 1
                        ws.cell(row=i, column=date+1).value = attendance

                if entered_year == end_year and entered_month_index == end_month_index and entered_date == end_date:
                    break
                
                if prev_date != date and employee_attendance_flag == 0:
                    attendance = 'A'
                    ws.cell(row=i, column=date+1).value = attendance

                prev_date = date
                employee_attendance_flag = 0

    query_database_calculate_wo_td(ws)

    wb.save(employee_name + ".xlsx")

    count = local_count
    print("File generated!")

